"""Dump xlsx to markdown mirror — one H2 per sheet, table per sheet content."""
import openpyxl, sys, pathlib, re

def cell(v):
    if v is None: return ''
    s = str(v).replace('|', '\\|').replace('\n', '<br>').strip()
    return s

def dump(xlsx_path, md_path, doc_title):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    out = []
    out.append(f"# {doc_title}\n")
    out.append(f"> **Source of Truth Mirror** — 自動 dump 自 [`{pathlib.Path(xlsx_path).name}`](../../{pathlib.Path(xlsx_path).name})\n")
    out.append(f"> **Generated**: 2026-05-27 from Excel ({len(wb.sheetnames)} sheets)\n")
    out.append(f"> **Status**: read-only mirror; 業主編輯只改 xlsx，docs 自動同步\n")
    out.append(f"> **D4 雙存治理**: docs 內引用走 markdown anchor (`#sheet-NN-name`)，不直接引 xlsx\n\n")
    out.append("## Table of Contents\n")
    for sn in wb.sheetnames:
        anchor = re.sub(r'[^a-z0-9]+', '-', sn.lower()).strip('-')
        out.append(f"- [{sn}](#{anchor})")
    out.append("")
    out.append("---\n")
    for sn in wb.sheetnames:
        ws = wb[sn]
        out.append(f"## {sn}\n")
        rows = [[cell(c) for c in row] for row in ws.iter_rows(values_only=True)]
        rows = [r for r in rows if any(c for c in r)]
        if not rows:
            out.append("_(empty sheet)_\n")
            continue
        # Try to detect header row: first row that looks like headers (non-empty, no merged)
        # Use first non-empty row as header
        header_idx = 0
        for i, r in enumerate(rows):
            if sum(1 for c in r if c) >= 2:
                header_idx = i
                break
        # Lead description rows (before header) as quoted text
        for i in range(header_idx):
            line = ' '.join(c for c in rows[i] if c)
            if line:
                out.append(f"> {line}\n")
        # Build table
        header = rows[header_idx]
        # Pad shorter rows
        ncol = max(len(r) for r in rows[header_idx:])
        header = (header + [''] * ncol)[:ncol]
        # Filter trailing empty header cells
        last_used = ncol
        for i in range(ncol-1, -1, -1):
            if header[i] or any(r[i] if i < len(r) else '' for r in rows[header_idx+1:]):
                last_used = i + 1
                break
        header = header[:last_used]
        # Use 'col_N' for empty headers
        header = [h if h else f"col_{i+1}" for i,h in enumerate(header)]
        out.append('| ' + ' | '.join(header) + ' |')
        out.append('|' + '|'.join(['---']*len(header)) + '|')
        for r in rows[header_idx+1:]:
            r = (r + [''] * len(header))[:len(header)]
            out.append('| ' + ' | '.join(r) + ' |')
        out.append("")
    pathlib.Path(md_path).write_text('\n'.join(out), encoding='utf-8')
    return len(wb.sheetnames)

n1 = dump('01-workorder-erp-final-spec-20260520.xlsx',
          'docs/_source/01-workorder-erp.md',
          'WorkOrder ERP Final Spec (2026-05-20)')
n2 = dump('02-ai-chatbot-sync-final-spec-20260520.xlsx',
          'docs/_source/02-ai-chatbot-sync.md',
          'AI Chatbot + WorkOrder Sync Final Spec (2026-05-20)')
print(f'dumped: ERP={n1} sheets, Chatbot/Sync={n2} sheets')
